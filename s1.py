import logging
import json
import os
import time
from datetime import datetime, timedelta
from telegram import Update, Bot
from telegram.ext import Updater, CommandHandler, CallbackContext, JobQueue
from binance.client import Client
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configurations
TOKEN = "7637023247:AAG_utVTC0rXyfute9xsBdh-IrTUE3432o8"
BINANCE_API_KEY = "aVim4czsoOzuLxk0CsEsV0JwE58OX90GRD8OvDfT8xH2nfSEC0mMnMCVrwgFcSEi"
BINANCE_API_SECRET = "rIQ2LLUtYWBcXt5FiMIHuXeeDJqeREbvw8r9NlTJ83gveSAvpSMqd1NBoQjAodC4"
CHAT_ID = 7662080576
LOG_FILE = "bot_log.json"
PATTERN_LOG_FILE = "pattern_log.txt"
DEBUG_LOG_FILE = "debug_log.txt"
EXCEL_FILE = "pivots.xlsx"

# Setup Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ensure log files exist
for file in [LOG_FILE, PATTERN_LOG_FILE, DEBUG_LOG_FILE]:
    if not os.path.exists(file):
        with open(file, "w", encoding="utf-8") as f:
            f.write("=== Log Initialized ===\n")

# Initialize Binance Client
binance_client = Client(BINANCE_API_KEY, BINANCE_API_SECRET)

class PivotData:
    def __init__(self):
        self.detected_pivots = []  
        self.user_provided_pivots = []  
        self.MIN_PRICE_CHANGE = 0.004  # Tăng từ 0.002 lên 0.004 vì timeframe lớn hơn
        self.MIN_PIVOT_DISTANCE = 2    # Giảm từ 3 xuống 2 vì khoảng thời gian giữa các nến dài hơn
        self.TREND_WINDOW = 5          # Giảm từ 10 xuống 5 vì timeframe lớn hơn
        self.MAX_PIVOTS = 15  
        self.last_sync_time = datetime.now()
        self.CONFIRMATION_CANDLES = 2   # Giảm từ 3 xuống 2 vì timeframe lớn hơn
        self.pending_pivots = []

    def add_price_data(self, price_data: dict):
        """Thêm dữ liệu giá mới vào price_history"""
        if not hasattr(self, 'price_history'):
            self.price_history = []
        
        self.price_history.append(price_data)
        if len(self.price_history) > 100:  # Giữ 100 điểm dữ liệu gần nhất
            self.price_history.pop(0)
        save_log(f"Đã thêm dữ liệu giá: {price_data}", DEBUG_LOG_FILE)
    
    def get_pivot_support_resistance(self, lookback: int = 20) -> dict:
        """
        Tính toán các mức hỗ trợ và kháng cự dựa trên pivot points
        Returns:
            Dict chứa các mức S/R và độ mạnh của chúng
        """
        try:
            if not hasattr(self, 'price_history') or len(self.price_history) < lookback:
                save_log(f"Không đủ dữ liệu để tính S/R (cần {lookback})", DEBUG_LOG_FILE)
                return {}

            # Lấy dữ liệu trong khoảng lookback
            recent_data = self.price_history[-lookback:]
            
            # Tính PP (Pivot Point)
            highs = [x['high'] for x in recent_data]
            lows = [x['low'] for x in recent_data]
            closes = [x['price'] for x in recent_data]
            
            pp = (max(highs) + min(lows) + closes[-1]) / 3
            
            # Tính các mức S/R
            r3 = pp + (max(highs) - min(lows))
            r2 = pp + (max(highs) - min(lows)) * 0.618  # Fibonacci ratio
            r1 = 2 * pp - min(lows)
            
            s1 = 2 * pp - max(highs)
            s2 = pp - (max(highs) - min(lows)) * 0.618
            s3 = pp - (max(highs) - min(lows))
            
            # Tính độ mạnh của mỗi mức
            def calculate_strength(level):
                touches = sum(1 for price in closes if abs(price - level) / level < 0.001)
                return min(touches / lookback * 100, 100)  # Độ mạnh tối đa 100%
            
            levels = {
                "R3": {"price": r3, "strength": calculate_strength(r3)},
                "R2": {"price": r2, "strength": calculate_strength(r2)},
                "R1": {"price": r1, "strength": calculate_strength(r1)},
                "PP": {"price": pp, "strength": calculate_strength(pp)},
                "S1": {"price": s1, "strength": calculate_strength(s1)},
                "S2": {"price": s2, "strength": calculate_strength(s2)},
                "S3": {"price": s3, "strength": calculate_strength(s3)}
            }
            
            save_log(f"Đã tính toán mức S/R: {levels}", DEBUG_LOG_FILE)
            return levels

        except Exception as e:
            save_log(f"Lỗi tính S/R: {str(e)}", DEBUG_LOG_FILE)
            return {}
    
    def improve_pivot_detection(self, price: float, time: str) -> tuple[bool, str]:
        """
        Cải thiện logic xác định pivot bằng cách kết hợp với mức S/R
        Returns:
            (is_pivot, pivot_type)
        """
        try:
            # Lấy mức S/R
            support_resistance = self.get_pivot_support_resistance()
            if not support_resistance:
                return False, ""

            # Kiểm tra xem giá có gần mức S/R nào không
            MIN_DISTANCE = 0.001  # 0.1% cho phép dao động
            
            for level_name, level_data in support_resistance.items():
                level_price = level_data["price"]
                level_strength = level_data["strength"]
                
                price_diff = abs(price - level_price) / level_price
                
                if price_diff <= MIN_DISTANCE:
                    # Giá chạm mức S/R
                    if level_strength >= 70:  # Mức S/R mạnh
                        if "R" in level_name:  # Mức kháng cự
                            save_log(f"Phát hiện pivot tại mức kháng cự {level_name}: ${price:,.2f}", DEBUG_LOG_FILE)
                            return True, "High"
                        elif "S" in level_name:  # Mức hỗ trợ
                            save_log(f"Phát hiện pivot tại mức hỗ trợ {level_name}: ${price:,.2f}", DEBUG_LOG_FILE)
                            return True, "Low"
            
            return False, ""

        except Exception as e:
            save_log(f"Lỗi cải thiện pivot: {str(e)}", DEBUG_LOG_FILE)
            return False, ""
    
    def analyze_market_trend(self, short_period: int = 10, medium_period: int = 20, long_period: int = 50) -> dict:
        """
        Phân tích xu hướng thị trường sử dụng nhiều chỉ báo
        Returns:
            Dict chứa kết quả phân tích
        """
        try:
            if not hasattr(self, 'price_history') or len(self.price_history) < long_period:
                save_log(f"Không đủ dữ liệu để phân tích (cần {long_period})", DEBUG_LOG_FILE)
                return {}

            prices = [x['price'] for x in self.price_history]
            
            # Tính MA các chu kỳ
            def calculate_ma(period):
                if len(prices) < period:
                    return None
                return sum(prices[-period:]) / period
            
            short_ma = calculate_ma(short_period)
            medium_ma = calculate_ma(medium_period)
            long_ma = calculate_ma(long_period)
            
            # Tính RSI
            def calculate_rsi(period=14):
                if len(prices) < period + 1:
                    return None
                    
                deltas = [prices[i+1] - prices[i] for i in range(len(prices)-1)]
                gains = [d if d > 0 else 0 for d in deltas]
                losses = [-d if d < 0 else 0 for d in deltas]
                
                avg_gain = sum(gains[-period:]) / period
                avg_loss = sum(losses[-period:]) / period
                
                if avg_loss == 0:
                    return 100
                
                rs = avg_gain / avg_loss
                rsi = 100 - (100 / (1 + rs))
                return rsi
                
            rsi = calculate_rsi()
            
            # Xác định xu hướng
            trend = "Unknown"
            strength = 0
            
            if short_ma and medium_ma and long_ma:
                if short_ma > medium_ma > long_ma:
                    trend = "Uptrend"
                    strength = min(((short_ma/long_ma - 1) * 100), 100)
                elif short_ma < medium_ma < long_ma:
                    trend = "Downtrend"
                    strength = min(((1 - short_ma/long_ma) * 100), 100)
                else:
                    trend = "Sideways"
                    strength = 0
                    
            # Tính volatility
            if len(prices) >= 20:
                recent_prices = prices[-20:]
                avg_price = sum(recent_prices) / len(recent_prices)
                volatility = sum([abs(p - avg_price) / avg_price for p in recent_prices]) / len(recent_prices) * 100
            else:
                volatility = None

            result = {
                "trend": trend,
                "strength": strength,
                "short_ma": short_ma,
                "medium_ma": medium_ma,
                "long_ma": long_ma,
                "rsi": rsi,
                "volatility": volatility
            }
            
            save_log(f"Kết quả phân tích xu hướng: {result}", DEBUG_LOG_FILE)
            return result

        except Exception as e:
            save_log(f"Lỗi phân tích xu hướng: {str(e)}", DEBUG_LOG_FILE)
            return {}
   
    def add_user_pivot(self, pivot_type: str, price: float, time: str) -> bool:
        """Thêm pivot từ lệnh /moc của user"""
        try:
            new_pivot = {
                "type": pivot_type.upper(),
                "price": price,
                "time": time,
                "source": "user"
            }
            self.user_provided_pivots.append(new_pivot)
            
            # Giới hạn số lượng pivot
            if len(self.user_provided_pivots) > self.MAX_PIVOTS:
                self.user_provided_pivots.pop(0)
            
            save_log(f"User pivot added: {pivot_type} at {time} price: ${price}", DEBUG_LOG_FILE)
            return True
        except Exception as e:
            save_log(f"Error adding user pivot: {str(e)}", DEBUG_LOG_FILE)
            return False

    def add_detected_pivot(self, price: float, price_type: str) -> bool:
        """Thêm pivot từ hệ thống tự động phát hiện"""
        try:
            # Kiểm tra điều kiện thêm pivot
            if not self._can_add_pivot(price):
                return False

            # Xác định loại pivot
            pivot_type = self._determine_pivot_type(price, price_type)
            if not pivot_type:
                return False

            # Tạo pivot mới
            new_pivot = {
                "type": pivot_type,
                "price": price,
                "time": datetime.now().strftime("%H:%M"),
                "source": "system"
            }
            self.detected_pivots.append(new_pivot)

            # Giới hạn số lượng pivot
            if len(self.detected_pivots) > self.MAX_PIVOTS:
                self.detected_pivots.pop(0)

            save_log(f"Detected pivot: {pivot_type} at {new_pivot['time']} price: ${price}", DEBUG_LOG_FILE)
            return True
        except Exception as e:
            save_log(f"Error adding detected pivot: {str(e)}", DEBUG_LOG_FILE)
            return False

    def _can_add_pivot(self, price: float) -> bool:
        """Kiểm tra các điều kiện để thêm pivot mới"""
        try:
            combined_pivots = self.get_all_pivots()
            if not combined_pivots:
                save_log("Không có pivot trước đó, cho phép thêm mới", DEBUG_LOG_FILE)
                return True

            last_pivot = combined_pivots[-1]
            save_log(f"Kiểm tra điều kiện thêm pivot mới: ${price:,.2f} vs Last=${last_pivot['price']:,.2f}", DEBUG_LOG_FILE)

            # Kiểm tra khoảng cách thời gian
            last_time = datetime.strptime(last_pivot["time"], "%H:%M")
            current_time = datetime.now()
            time_diff = (current_time - last_time).total_seconds() / 1800  # Đổi sang số nến 30m
            
            save_log(f"Khoảng cách thời gian: {time_diff:.1f} nến 30m", DEBUG_LOG_FILE)
            
            if time_diff < self.MIN_PIVOT_DISTANCE:
                save_log(f"Từ chối pivot: Quá gần thời gian ({time_diff:.1f} < {self.MIN_PIVOT_DISTANCE} nến)", DEBUG_LOG_FILE)
                return False

            # Kiểm tra biến động giá
            price_change = abs(price - last_pivot["price"]) / last_pivot["price"]
            save_log(f"Biên độ giá: {price_change:.2%}", DEBUG_LOG_FILE)
            
            if price_change < self.MIN_PRICE_CHANGE:
                save_log(f"Từ chối pivot: Biên độ quá nhỏ ({price_change:.2%} < {self.MIN_PRICE_CHANGE:.2%})", DEBUG_LOG_FILE)
                return False

            # Kiểm tra xu hướng nếu có đủ dữ liệu
            if len(self.price_history) >= self.TREND_WINDOW:
                trend = self._calculate_trend([x["price"] for x in self.price_history[-self.TREND_WINDOW:]])
                save_log(f"Xu hướng hiện tại: {trend:+d}", DEBUG_LOG_FILE)
                
                # Thêm logic xử lý dựa trên xu hướng
                if trend > 0:  # Xu hướng tăng
                    if price < last_pivot["price"] and last_pivot["type"] in ["HH", "LH"]:
                        save_log("Từ chối pivot: Giá thấp hơn trong xu hướng tăng", DEBUG_LOG_FILE)
                        return False
                elif trend < 0:  # Xu hướng giảm
                    if price > last_pivot["price"] and last_pivot["type"] in ["LL", "HL"]:
                        save_log("Từ chối pivot: Giá cao hơn trong xu hướng giảm", DEBUG_LOG_FILE)
                        return False

            save_log("Chấp nhận pivot mới: Đã thỏa mãn tất cả điều kiện", DEBUG_LOG_FILE)
            return True

        except Exception as e:
            save_log(f"Lỗi kiểm tra điều kiện thêm pivot: {str(e)}", DEBUG_LOG_FILE)
            return False
       
    def _determine_pivot_type(self, price: float, price_type: str) -> str:
        """Xác định loại pivot dựa trên xu hướng và cấu trúc giá"""
        try:
            all_pivots = self.get_all_pivots()
            if len(all_pivots) < 5:
                return self._determine_initial_pivot_type(price, price_type, all_pivots)
                
            last_5_pivots = [p["price"] for p in all_pivots[-5:]]
            if len(last_5_pivots) < 5:
                return self._determine_initial_pivot_type(price, price_type, all_pivots)
                
            a, b, c, d, e = last_5_pivots
            trend = self._calculate_trend(last_5_pivots)
            
            if price_type == "H":
                if trend > 0:  # Xu hướng tăng
                    if price > max(last_5_pivots):
                        return "HH"
                    elif c > b and c > d and price > c:
                        return "HH"
                    else:
                        return "LH"
                else:  # Xu hướng giảm
                    if price < min(last_5_pivots):
                        return "LH"
                    else:
                        return self._verify_lower_high(price, last_5_pivots)
            else:  # price_type == "L"
                if trend < 0:  # Xu hướng giảm
                    if price < min(last_5_pivots):
                        return "LL"
                    elif c < b and c < d and price < c:
                        return "LL"
                    else:
                        return "HL"
                else:  # Xu hướng tăng
                    if price > max(last_5_pivots):
                        return "HL"
                    else:
                        return self._verify_higher_low(price, last_5_pivots)
                        
        except Exception as e:
            save_log(f"Error determining pivot type: {str(e)}", DEBUG_LOG_FILE)
            return None

    def _determine_initial_pivot_type(self, price: float, price_type: str, pivots: list) -> str:
        """Xác định loại pivot khi có ít dữ liệu"""
        if not pivots:
            return "HH" if price_type == "H" else "LL"
        
        last_pivot = pivots[-1]
        if price_type == "H":
            return "HH" if price > last_pivot["price"] else "LH"
        else:
            return "LL" if price < last_pivot["price"] else "HL"

    def _verify_lower_high(self, price: float, prices: list) -> str:
        """Xác minh điểm LH"""
        higher_prices = [p for p in prices if p > price]
        if not higher_prices:
            return None
        avg_high = sum(higher_prices) / len(higher_prices)
        return "LH" if price < avg_high else None

    def _verify_higher_low(self, price: float, prices: list) -> str:
        """Xác minh điểm HL"""
        lower_prices = [p for p in prices if p < price]
        if not lower_prices:
            return None
        avg_low = sum(lower_prices) / len(lower_prices)
        return "HL" if price > avg_low else None

    def get_all_pivots(self) -> list:
        """Lấy tất cả pivot đã sắp xếp theo thời gian"""
        all_pivots = self.user_provided_pivots + self.detected_pivots
        all_pivots.sort(key=lambda x: datetime.strptime(x["time"], "%H:%M"))
        return all_pivots

    def clear_all(self):
        """Xóa tất cả pivot points"""
        self.detected_pivots.clear()
        self.user_provided_pivots.clear()
        save_log("All pivot points cleared", DEBUG_LOG_FILE)

    def get_recent_pivots(self, count: int = 5) -> list:
        """Lấy số lượng pivot gần nhất"""
        all_pivots = self.get_all_pivots()
        return all_pivots[-count:] if all_pivots else []

    def check_pattern(self) -> tuple[bool, str]:
        """Kiểm tra mẫu hình và trả về (có_mẫu_hình, tên_mẫu_hình)"""
        patterns = {
            "bullish_reversal": [
                ["HH", "HL", "HH", "HL", "HH"],
                ["LH", "HL", "HH", "HL", "HH"],
                ["HH", "HH", "HH"],
                ["HH", "HL", "HH", "HH"]
            ],
            "bearish_reversal": [
                ["LL", "LL", "LH", "LL"],
                ["LL", "LH", "LL", "LH", "LL"],
                ["LL", "LL", "LL"],
                ["LL", "LH", "LL", "LH", "LL"],
                ["LL", "LH", "LL"]
            ]
        }

        last_pivots = [p["type"] for p in self.get_all_pivots()]
        for pattern_name, sequences in patterns.items():
            for sequence in sequences:
                if len(last_pivots) >= len(sequence):
                    if last_pivots[-len(sequence):] == sequence:
                        save_log(f"Pattern found: {pattern_name} ({','.join(sequence)})", PATTERN_LOG_FILE)
                        return True, pattern_name
        return False, ""
        
    def _calculate_trend(self, prices: list) -> int:
        """Tính toán xu hướng dựa trên giá"""
        if len(prices) < 2:
            return 0
            
        changes = [prices[i] - prices[i-1] for i in range(1, len(prices))]
        up_moves = sum(1 for x in changes if x > 0)
        down_moves = sum(1 for x in changes if x < 0)
        
        if up_moves > down_moves:
            return 1
        elif down_moves > up_moves:
            return -1
        return 0    
        
    def remove_pivot(self, pivot_to_remove):
        """Xóa một pivot cụ thể"""
        try:
            if pivot_to_remove["source"] == "system":
                self.detected_pivots = [p for p in self.detected_pivots if p != pivot_to_remove]
            else:
                self.user_provided_pivots = [p for p in self.user_provided_pivots if p != pivot_to_remove]
            save_log(f"Đã xóa pivot: {pivot_to_remove}", DEBUG_LOG_FILE)
        except Exception as e:
            save_log(f"Lỗi khi xóa pivot: {str(e)}", DEBUG_LOG_FILE)  
    
    def add_pending_pivot(self, price: float, price_type: str, current_time: str):
        """Thêm một pivot vào danh sách chờ xác nhận"""
        try:
            pending_pivot = {
                "type": price_type,
                "price": price,
                "time": current_time,
                "confirmation_candles": 0,
                "higher_prices": 0,  # Số nến có giá cao hơn
                "lower_prices": 0,   # Số nến có giá thấp hơn
                "source": "system"
            }
            self.pending_pivots.append(pending_pivot)
            save_log(f"Thêm pivot chờ xác nhận: {price_type} tại ${price:,.2f}", DEBUG_LOG_FILE)
            return True
        except Exception as e:
            save_log(f"Lỗi khi thêm pending pivot: {str(e)}", DEBUG_LOG_FILE)
            return False

    def validate_pending_pivots(self, current_high: float, current_low: float) -> list:
        """Kiểm tra và xác nhận các pivot đang chờ"""
        try:
            confirmed_pivots = []
            remaining_pivots = []

            for pivot in self.pending_pivots:
                save_log(f"Đánh giá pivot: {pivot['type']} tại ${pivot['price']:,.2f} ({pivot['confirmation_candles']}/{self.CONFIRMATION_CANDLES} nến)", DEBUG_LOG_FILE)
                
                pivot["confirmation_candles"] += 1
                
                # Cập nhật số lượng nến cao/thấp hơn
                if pivot["type"] in ["H", "HH", "LH"]:  # Đỉnh
                    if current_high > pivot["price"]:
                        pivot["higher_prices"] += 1
                        save_log(f"Nến hiện tại cao hơn pivot đỉnh (${current_high:,.2f} > ${pivot['price']:,.2f})", DEBUG_LOG_FILE)
                    elif current_high < pivot["price"]:
                        pivot["lower_prices"] += 1
                        save_log(f"Nến hiện tại thấp hơn pivot đỉnh (${current_high:,.2f} < ${pivot['price']:,.2f})", DEBUG_LOG_FILE)
                else:  # Đáy
                    if current_low < pivot["price"]:
                        pivot["lower_prices"] += 1
                        save_log(f"Nến hiện tại thấp hơn pivot đáy (${current_low:,.2f} < ${pivot['price']:,.2f})", DEBUG_LOG_FILE)
                    elif current_low > pivot["price"]:
                        pivot["higher_prices"] += 1
                        save_log(f"Nến hiện tại cao hơn pivot đáy (${current_low:,.2f} > ${pivot['price']:,.2f})", DEBUG_LOG_FILE)

                # Kiểm tra đủ số nến xác nhận
                if pivot["confirmation_candles"] >= self.CONFIRMATION_CANDLES:
                    save_log(f"Đủ số nến xác nhận ({pivot['confirmation_candles']}/{self.CONFIRMATION_CANDLES})", DEBUG_LOG_FILE)
                    
                    if pivot["type"] in ["H", "HH", "LH"]:  # Xác nhận đỉnh
                        if pivot["lower_prices"] >= 2:  # Ít nhất 2 nến thấp hơn
                            save_log(f"Xác nhận đỉnh: {pivot['lower_prices']}/{self.CONFIRMATION_CANDLES} nến thấp hơn", DEBUG_LOG_FILE)
                            confirmed_pivots.append(pivot)
                        else:
                            save_log(f"Từ chối đỉnh: chỉ có {pivot['lower_prices']}/{self.CONFIRMATION_CANDLES} nến thấp hơn", DEBUG_LOG_FILE)
                    else:  # Xác nhận đáy
                        if pivot["higher_prices"] >= 2:  # Ít nhất 2 nến cao hơn
                            save_log(f"Xác nhận đáy: {pivot['higher_prices']}/{self.CONFIRMATION_CANDLES} nến cao hơn", DEBUG_LOG_FILE)
                            confirmed_pivots.append(pivot)
                        else:
                            save_log(f"Từ chối đáy: chỉ có {pivot['higher_prices']}/{self.CONFIRMATION_CANDLES} nến cao hơn", DEBUG_LOG_FILE)
                else:
                    remaining_pivots.append(pivot)
                    save_log(f"Chưa đủ nến xác nhận ({pivot['confirmation_candles']}/{self.CONFIRMATION_CANDLES}), giữ lại trong danh sách chờ", DEBUG_LOG_FILE)

                self.pending_pivots = remaining_pivots
                return confirmed_pivots

            except Exception as e:
                save_log(f"Lỗi khi xác nhận pending pivots: {str(e)}", DEBUG_LOG_FILE)
                return []
    
# Initialize PivotData instance
pivot_data = PivotData()

def save_log(log_message, filename):
    """ Save log messages to a text file """
    with open(filename, "a", encoding="utf-8") as f:
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} [INFO] - {log_message}\n")

def save_to_excel():
    """ 
    Lưu dữ liệu pivot vào file Excel với các cải tiến:
    - Phân biệt pivot từ user và hệ thống
    - Thêm biểu đồ candlestick
    - Cải thiện định dạng và bố cục
    """
    try:
        all_pivots = pivot_data.get_all_pivots()
        if not all_pivots:
            save_log("No pivot data to save", DEBUG_LOG_FILE)
            return
        
        wb = Workbook()
        # Tạo worksheet cho pivot points
        ws_pivot = wb.active
        ws_pivot.title = "Pivot Points"
        
        # Định dạng tiêu đề
        headers = ["Time", "Type", "Price", "Source", "Change %", "Trend"]
        for col, header in enumerate(headers, 1):
            cell = ws_pivot.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            ws_pivot.column_dimensions[get_column_letter(col)].width = 15

        # Thêm dữ liệu với định dạng màu và tính toán bổ sung
        prev_price = None
        trend = "N/A"
        
        for idx, pivot in enumerate(all_pivots, 2):
            # Thêm thông tin cơ bản
            ws_pivot.cell(row=idx, column=1, value=pivot["time"])
            ws_pivot.cell(row=idx, column=2, value=pivot["type"])
            ws_pivot.cell(row=idx, column=3, value=pivot["price"])
            ws_pivot.cell(row=idx, column=4, value=pivot["source"])
            
            # Tính % thay đổi và xu hướng
            if prev_price:
                change = ((pivot["price"] - prev_price) / prev_price) * 100
                ws_pivot.cell(row=idx, column=5, value=f"{change:+.2f}%")
                
                # Xác định xu hướng
                if change > 0:
                    trend = "↗ Tăng"
                    cell_color = "00FF00"  # Màu xanh lá
                elif change < 0:
                    trend = "↘ Giảm"
                    cell_color = "FF0000"  # Màu đỏ
                else:
                    trend = "→ Đi ngang"
                    cell_color = "FFFF00"  # Màu vàng
                
                # Thêm xu hướng và định dạng màu
                trend_cell = ws_pivot.cell(row=idx, column=6, value=trend)
                trend_cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
                
            prev_price = pivot["price"]
            
            # Định dạng các ô
            for col in range(1, 7):
                cell = ws_pivot.cell(row=idx, column=col)
                cell.alignment = Alignment(horizontal="center")
                
                # Thêm màu nền cho các pivot từ user
                if pivot["source"] == "user":
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

        # Tạo biểu đồ
        chart = LineChart()
        chart.title = "Pivot Points Analysis"
        chart.style = 13
        chart.height = 15
        chart.width = 30
        
        # Dữ liệu cho biểu đồ
        data = Reference(ws_pivot, min_col=3, min_row=1, max_row=len(all_pivots) + 1)
        categories = Reference(ws_pivot, min_col=1, min_row=2, max_row=len(all_pivots) + 1)
        
        # Thêm series và định dạng
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Định dạng trục
        chart.x_axis.title = "Time"
        chart.y_axis.title = "Price (USD)"
        chart.x_axis.tickLblSkip = 2
        
        # Thêm các điểm đánh dấu
        series = chart.series[0]
        series.marker.symbol = "circle"
        series.marker.size = 8
        
        # Thêm biểu đồ vào worksheet
        ws_pivot.add_chart(chart, "H2")
        
        # Thêm thông tin tổng hợp
        summary_row = len(all_pivots) + 4
        ws_pivot.cell(row=summary_row, column=1, value="Thống kê:")
        ws_pivot.cell(row=summary_row + 1, column=1, value="Tổng số pivot:")
        ws_pivot.cell(row=summary_row + 1, column=2, value=len(all_pivots))
        ws_pivot.cell(row=summary_row + 2, column=1, value="Pivot từ user:")
        ws_pivot.cell(row=summary_row + 2, column=2, value=len([p for p in all_pivots if p["source"] == "user"]))
        ws_pivot.cell(row=summary_row + 3, column=1, value="Pivot từ hệ thống:")
        ws_pivot.cell(row=summary_row + 3, column=2, value=len([p for p in all_pivots if p["source"] == "system"]))
        
        # Lưu file
        wb.save(EXCEL_FILE)
        save_log(f"Pivot data saved to Excel with {len(all_pivots)} points", DEBUG_LOG_FILE)
        
    except Exception as e:
        error_msg = f"Error saving Excel file: {str(e)}"
        save_log(error_msg, DEBUG_LOG_FILE)
        logger.error(error_msg)
    
def get_binance_price(context: CallbackContext):
    try:
        # Thay đổi interval từ "5m" sang "30m"
        klines = binance_client.futures_klines(symbol="BTCUSDT", interval="30m", limit=2)
        last_candle = klines[-2]  # Ensure we get the closed candle
        high_price = float(last_candle[2])
        low_price = float(last_candle[3])
        close_price = float(last_candle[4])
        
        price_data = {
            "high": high_price,
            "low": low_price,
            "price": close_price
        }
        pivot_data.add_price_data(price_data)
        
        save_log(f"Thu thập dữ liệu nến 30m: Cao nhất = {high_price}, Thấp nhất = {low_price}", DEBUG_LOG_FILE)
        
        detect_pivot(high_price, "H")
        detect_pivot(low_price, "L")
        save_to_excel()
    except Exception as e:
        logger.error(f"Binance API Error: {e}")
        save_log(f"Binance API Error: {e}", DEBUG_LOG_FILE)
        
def schedule_next_run(job_queue):
    try:
        # lên lịch chạy khi chẵn 30p
        now = datetime.now()
        next_run = now.replace(second=0, microsecond=0) + timedelta(minutes=(30 - now.minute % 30))
        delay = (next_run - now).total_seconds()
        
        save_log(f"Lên lịch chạy vào {next_run.strftime('%Y-%m-%d %H:%M:%S')}", DEBUG_LOG_FILE)
        # Thay đổi interval từ 300 (5 phút) sang 1800 (30 phút)
        job_queue.run_repeating(get_binance_price, interval=1800, first=delay)
    except Exception as e:
        logger.error(f"Error scheduling next run: {e}")
        save_log(f"Error scheduling next run: {e}", DEBUG_LOG_FILE)

def detect_pivot(price, price_type):
    """Phát hiện và xử lý pivot points"""
    try:
        current_time = datetime.now().strftime("%H:%M")
        
        # Validate các pivot đang chờ
        if hasattr(pivot_data, 'price_history') and len(pivot_data.price_history) > 0:
            current_candle = pivot_data.price_history[-1]
            
            # So sánh biên độ TRƯỚC khi xác nhận
            all_pivots = pivot_data.get_all_pivots()
            if all_pivots:
                current_price_change = abs(price - all_pivots[-1]["price"]) / price
                last_pivot = all_pivots[-1]
                last_price_change = abs(last_pivot["price"] - (all_pivots[-2]["price"] if len(all_pivots) > 1 else last_pivot["price"])) / last_pivot["price"]
                
                save_log(f"So sánh biên độ trước xác nhận: Hiện tại {current_price_change:.2%} vs Cũ {last_price_change:.2%}", DEBUG_LOG_FILE)
                
                # Nếu biên độ nhỏ hơn, bỏ qua pivot này
                if current_price_change <= last_price_change:
                    save_log(f"Bỏ qua pivot do biên độ nhỏ hơn (${price:,.2f})", DEBUG_LOG_FILE)
                    return

            # Sau khi kiểm tra biên độ, tiến hành xác nhận pivot
            confirmed_pivots = pivot_data.validate_pending_pivots(
                current_high=current_candle['high'],
                current_low=current_candle['low']
            )

            # Thêm các pivot đã được xác nhận
            for confirmed_pivot in confirmed_pivots:
                # Kiểm tra lại biên độ một lần nữa trước khi thêm
                if all_pivots:
                    confirmed_price_change = abs(confirmed_pivot["price"] - all_pivots[-1]["price"]) / confirmed_pivot["price"]
                    if confirmed_price_change <= last_price_change:
                        save_log(f"Bỏ qua pivot đã xác nhận do biên độ nhỏ hơn (${confirmed_pivot['price']:,.2f})", DEBUG_LOG_FILE)
                        continue

                if pivot_data.add_detected_pivot(confirmed_pivot["price"], confirmed_pivot["type"]):
                    save_log(f"Pivot đã được xác nhận và thêm vào: {confirmed_pivot['type']} - ${confirmed_pivot['price']:,.2f}", DEBUG_LOG_FILE)
                    save_to_excel()

                    # Kiểm tra mẫu hình cho pivot mới xác nhận
                    has_pattern, pattern_name = pivot_data.check_pattern()
                    if has_pattern:
                        recent_pivots = pivot_data.get_recent_pivots(5)
                        message = _create_alert_message(pattern_name, price, recent_pivots)
                        send_alert(message)

        # Xem xét thêm pivot mới vào danh sách chờ
        all_pivots = pivot_data.get_all_pivots()
        
        # Kiểm tra pivot trong cùng thời điểm
        for pivot in all_pivots:
            if pivot["time"] == current_time:
                save_log(f"Đã có pivot tại {current_time}, kiểm tra biên độ", DEBUG_LOG_FILE)
                return

        # Kiểm tra điều kiện thêm pivot mới
        if not pivot_data._can_add_pivot(price):
            return

        # Xác định loại pivot
        pivot_type = pivot_data._determine_pivot_type(price, price_type)
        if not pivot_type:
            return

        # Thêm vào danh sách chờ xác nhận
        if pivot_data.add_pending_pivot(price, pivot_type, current_time):
            save_log(f"Đã thêm pivot vào danh sách chờ: {pivot_type} - Giá: ${price:,.2f} vào lúc {current_time}", DEBUG_LOG_FILE)
            
    except Exception as e:
        error_msg = f"Lỗi trong quá trình phát hiện pivot: {str(e)}"
        save_log(error_msg, DEBUG_LOG_FILE)
        logger.error(error_msg)

def _create_alert_message(pattern_name, current_price, recent_pivots):
    """Tạo thông báo chi tiết khi phát hiện mẫu hình"""
    vietnam_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Xác định loại mẫu hình và biểu tượng
    if "bullish" in pattern_name.lower():
        pattern_symbol = "🟢"
        direction = "tăng"
    else:
        pattern_symbol = "🔴"
        direction = "giảm"
        
    message = (
        f"{pattern_symbol} CẢNH BÁO MẪU HÌNH {direction.upper()} - {vietnam_time}\n\n"
        f"Giá hiện tại: ${current_price:,.2f}\n"
        f"Mẫu hình: {pattern_name}\n\n"
        f"5 pivot gần nhất:\n"
    )
    
    # Thêm thông tin về 5 pivot gần nhất
    for i, pivot in enumerate(recent_pivots[::-1], 1):
        message += f"{i}. {pivot['type']}: ${pivot['price']:,.2f} ({pivot['time']})\n"
        
    return message

def send_alert(message):
    """Gửi cảnh báo qua Telegram với thông tin chi tiết"""
    try:
        bot = Bot(token=TOKEN)
        bot.send_message(
            chat_id=CHAT_ID,
            text=message,
            parse_mode='HTML'
        )
        save_log("Đã gửi cảnh báo mẫu hình", DEBUG_LOG_FILE)
    except Exception as e:
        save_log(f"Lỗi gửi cảnh báo: {str(e)}", DEBUG_LOG_FILE)

def moc(update: Update, context: CallbackContext):
    """ Handles the /moc command to receive multiple pivot points and resets logic."""
    try:
        args = context.args
        
        logger.info(f"Received /moc command with args: {args}")
        save_log(f"Received /moc command with args: {args}", DEBUG_LOG_FILE)
        
        if len(args) < 4 or (len(args) - 1) % 3 != 0:
            update.message.reply_text("⚠️ Sai định dạng! Dùng: /moc btc lh 82000 13:30 hl 81000 14:00 hh 83000 14:30")
            return
        
        asset = args[0].lower()
        if asset != "btc":
            update.message.reply_text("⚠️ Chỉ hỗ trợ BTC! Ví dụ: /moc btc lh 82000 13:30 hl 81000 14:00 hh 83000 14:30")
            return
            
        # Xóa dữ liệu cũ
        pivot_data.clear_all()
        
        # Ghi nhận các mốc mới
        valid_pivots = []
        adjusted_times = []
        current_time = datetime.now()  # Lấy thời gian hiện tại
        
        # Kiểm tra thứ tự thời gian
        time_points = []
        for i in range(1, len(args), 3):
            try:
                time = args[i + 2].replace('h', ':')
                time_obj = datetime.strptime(time, "%H:%M")
                time_points.append(time_obj)
            except ValueError:
                continue

        if time_points:
            if time_points != sorted(time_points):
                update.message.reply_text("⚠️ Các mốc thời gian phải được nhập theo thứ tự tăng dần!")
                return
        
        for i in range(1, len(args), 3):
            pivot_type = args[i].upper()
            if pivot_type not in ["HH", "HL", "LH", "LL"]:
                update.message.reply_text(f"⚠️ Loại pivot không hợp lệ: {pivot_type}. Chỉ chấp nhận: HH, HL, LH, LL")
                return

            # Validate giá
            try:
                price = float(args[i + 1])
                if price <= 0:
                    update.message.reply_text(f"⚠️ Giá phải lớn hơn 0: {args[i + 1]}")
                    return
                if price > 500000:  # Giới hạn giá tối đa hợp lý cho BTC
                    update.message.reply_text(f"⚠️ Giá vượt quá giới hạn cho phép: {args[i + 1]}")
                    return
            except ValueError:
                update.message.reply_text(f"⚠️ Giá không hợp lệ: {args[i + 1]}")
                return

            # Validate và xử lý thời gian
            time = args[i + 2].replace('h', ':')
            try:
                time_obj = datetime.strptime(time, "%H:%M")
                
                # Làm tròn về mốc 30 phút gần nhất
                minutes = time_obj.minute
                if minutes % 30 != 0:
                    adjusted_minutes = 30 * (minutes // 30)
                    original_time = time
                    time = time_obj.replace(minute=adjusted_minutes).strftime("%H:%M")
                    adjusted_times.append((original_time, time))
                    save_log(f"Đã điều chỉnh thời gian từ {original_time} thành {time} cho phù hợp với timeframe 30m", DEBUG_LOG_FILE)
            except ValueError:
                update.message.reply_text(f"⚠️ Lỗi: Định dạng thời gian không đúng! Sử dụng HH:MM (ví dụ: 14:00, 14:30)")
                return

            # Thêm pivot mới
            if pivot_data.add_user_pivot(pivot_type, price, time):
                valid_pivots.append({"type": pivot_type, "price": price, "time": time})
            else:
                update.message.reply_text(f"⚠️ Không thể thêm pivot: {pivot_type} at {time}")
                return
        
        # Kiểm tra tính hợp lệ của chuỗi pivot
        if len(valid_pivots) >= 2:
            for i in range(1, len(valid_pivots)):
                curr_pivot = valid_pivots[i]
                prev_pivot = valid_pivots[i-1]
                
                save_log(f"Kiểm tra logic: {curr_pivot['type']} (${curr_pivot['price']}) vs {prev_pivot['type']} (${prev_pivot['price']})", DEBUG_LOG_FILE)
                
                # Logic kiểm tra mới
                if curr_pivot['type'] == "LH":
                    if prev_pivot['type'] == "LL":
                        # LH phải cao hơn LL trước đó
                        if curr_pivot['price'] <= prev_pivot['price']:
                            error_msg = f"⚠️ Lỗi logic: LH tại {curr_pivot['time']} phải có giá cao hơn LL trước đó!"
                            save_log(error_msg, DEBUG_LOG_FILE)
                            update.message.reply_text(error_msg)
                            return
                    elif prev_pivot['type'] == "HH":
                        # LH phải thấp hơn HH trước đó 
                        if curr_pivot['price'] >= prev_pivot['price']:
                            error_msg = f"⚠️ Lỗi logic: LH tại {curr_pivot['time']} phải có giá thấp hơn HH trước đó!"
                            save_log(error_msg, DEBUG_LOG_FILE)
                            update.message.reply_text(error_msg)
                            return
                        
                elif curr_pivot['type'] == "HL":
                    if prev_pivot['type'] in ["LH", "HH"]:
                        # HL phải thấp hơn đỉnh trước đó (LH hoặc HH)
                        if curr_pivot['price'] >= prev_pivot['price']:
                            error_msg = f"⚠️ Lỗi logic: HL tại {curr_pivot['time']} phải có giá thấp hơn {prev_pivot['type']} trước đó!"
                            save_log(error_msg, DEBUG_LOG_FILE)
                            update.message.reply_text(error_msg)
                            return
                    elif prev_pivot['type'] == "LL":
                        # HL phải cao hơn LL trước đó
                        if curr_pivot['price'] <= prev_pivot['price']:
                            error_msg = f"⚠️ Lỗi logic: HL tại {curr_pivot['time']} phải có giá cao hơn LL trước đó!"
                            save_log(error_msg, DEBUG_LOG_FILE)
                            update.message.reply_text(error_msg)
                            return
                        
                elif curr_pivot['type'] == "HH":
                    # HH luôn phải cao hơn pivot trước đó
                    if curr_pivot['price'] <= prev_pivot['price']:
                        error_msg = f"⚠️ Lỗi logic: HH tại {curr_pivot['time']} phải có giá cao hơn pivot trước đó!"
                        save_log(error_msg, DEBUG_LOG_FILE)
                        update.message.reply_text(error_msg)
                        return
                        
                elif curr_pivot['type'] == "LL":
                    # LL luôn phải thấp hơn pivot trước đó
                    if curr_pivot['price'] >= prev_pivot['price']:
                        error_msg = f"⚠️ Lỗi logic: LL tại {curr_pivot['time']} phải có giá thấp hơn pivot trước đó!"
                        save_log(error_msg, DEBUG_LOG_FILE)
                        update.message.reply_text(error_msg)
                        return
                        
                save_log(f"Pivot {curr_pivot['type']} hợp lệ", DEBUG_LOG_FILE)
        
        # Ghi đè dữ liệu vào pattern log
        with open(PATTERN_LOG_FILE, "w", encoding="utf-8") as f:
            f.write("=== Pattern Log Reset ===\n")

        save_log(f"User Pivots Updated: {pivot_data.user_provided_pivots}", LOG_FILE)
        save_log(f"User Pivots Updated: {pivot_data.user_provided_pivots}", PATTERN_LOG_FILE)
        save_to_excel()

        # Tạo phản hồi chi tiết cho người dùng
        response = "✅ Đã nhận các mốc:\n"
        for pivot in valid_pivots:
            response += f"• {pivot['type']} tại ${pivot['price']:,.2f} ({pivot['time']})\n"
        
        # Thêm thông báo về các điều chỉnh thời gian (nếu có)
        if adjusted_times:
            response += "\nℹ️ Đã điều chỉnh các mốc thời gian sau cho phù hợp với timeframe 30m:\n"
            for original, adjusted in adjusted_times:
                response += f"• {original} → {adjusted}\n"
            
        update.message.reply_text(response)
        logger.info(f"User Pivots Updated: {pivot_data.user_provided_pivots}")
        
    except Exception as e:
        error_msg = f"Lỗi xử lý lệnh /moc: {str(e)}"
        logger.error(error_msg)
        save_log(error_msg, DEBUG_LOG_FILE)
        update.message.reply_text(f"⚠️ Có lỗi xảy ra: {str(e)}")

def main():
    """ Main entry point to start the bot."""
    try:
        updater = Updater(TOKEN, use_context=True)
        dp = updater.dispatcher
        job_queue = updater.job_queue
    
        dp.add_handler(CommandHandler("moc", moc))
        
        schedule_next_run(job_queue)  # Schedule the first execution at the next 5-minute mark
        
        print("Bot is running...")
        logger.info("Bot started successfully.")
        updater.start_polling()
        updater.idle()
    except Exception as e:
        logger.error(f"Error in main: {e}")
        save_log(f"Error in main: {e}", DEBUG_LOG_FILE)

if __name__ == "__main__":
    main()
